from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import cv2.cv2 as cv2
import json


def boxCenter(chkp):
    return [(chkp[0] + chkp[2]) / 2, (chkp[1] + chkp[3]) / 2]


def ifSameRow(si, ti):
    c1 = boxCenter(si)
    c2 = boxCenter(ti)
    ss, se = si[1], si[3]
    ts, te = ti[1], ti[3]
    if te >= c1[1] >= ts:
        return 1
    if se >= c2[1] >= ss:
        return 1
    return 0


def ifSameCol(si, ti):
    delta = 1
    c1 = boxCenter(si)
    c2 = boxCenter(ti)
    ss, se = si[0], si[2]
    ts, te = ti[0], ti[2]
    if te + delta >= c1[0] >= ts - delta:
        return 1
    if se + delta >= c2[0] >= ss - delta:
        return 1
    return 0


def cal_chk_limits(chunks):
    x_min = min(chunks, key=lambda p: p["pos"][0])["pos"][0]
    x_max = max(chunks, key=lambda p: p["pos"][1])["pos"][1]
    y_min = min(chunks, key=lambda p: p["pos"][2])["pos"][2]
    y_max = max(chunks, key=lambda p: p["pos"][3])["pos"][3]
    return [x_min, x_max, y_min, y_max, x_max - x_min, y_max - y_min]


def findUpChunk(chunks):
    up, minL = 0, 100000
    for i in range(len(chunks)):
        if chunks[i]['pos'][1] < minL:
            up = i
            minL = chunks[i]['pos'][1]
    return up


def findLeftChunk(chunks):
    left, minL = 0, 100000
    for i in range(len(chunks)):
        if chunks[i]['pos'][0] < minL:
            left = i
            minL = chunks[i]['pos'][0]
    return left


def processSpecial(chunks):
    rcs = []
    r = None
    for chunk in chunks:
        if chunk['text'] == '千港元':
            rcs.append((chunk['start_row'], chunk['start_col']))
            r = chunk['start_row']
            chunks.remove(chunk)
    for chunk in chunks:
        if (chunk['start_row'] + 1, chunk['start_col']) in rcs:
            chunk['text'] += ' ' + '千港元'
    if r is not None:
        for chunk in chunks:
            if chunk['start_row'] < r:
                chunk['start_row'] += 1
                chunk['end_row'] += 1
        for chunk in chunks:
            if chunk['start_row'] >= r:
                chunk['start_row'] -= 1
                chunk['end_row'] -= 1
    return chunks


def chunk2Structure(chunks):
    rows = []
    r = 0
    while len(chunks) != 0:
        upId = findUpChunk(chunks)
        row = []
        flag = 0
        # print(chunks[upId]['pos'])
        for chunk in chunks:
            if ifSameRow(chunks[upId]['pos'], chunk['pos']):
                # if chunk['text'] == '千港元':
                #     flag = 1
                #     row.append(chunk.copy())
                #     continue
                chunk['start_row'] = r
                chunk['end_row'] = r
                row.append(chunk.copy())
                # print(row)
        row = sorted(row, key=lambda x: x['pos'][0])
        for i in row:
            chunks.remove(i)
        if flag == 0:
            rows.append(row)
            r += 1
        # else:
        #     for item in rows[r-1]:

    chunks = []
    n_row = len(rows)
    for i in range(n_row):
        chunks.extend(rows[i])
    cols = []
    c = 0
    while len(chunks) != 0:
        leftid = findLeftChunk(chunks)
        # print('left:',chunks[leftid])
        col = []
        specialPos = []
        for chunk in chunks:
            if ifSameCol(chunks[leftid]['pos'], chunk['pos']):
                # if chunk['text'] == '千港元':
                #     specialPos.append((chunk['start_row'],c))
                chunk['start_col'] = c
                chunk['end_col'] = c
                col.append(chunk.copy())
        col = sorted(col, key=lambda x: x['pos'][2])
        for i in col:
            chunks.remove(i)
        cols.append(col)
        c += 1

    chunks = []
    n_col = len(cols)
    for i in range(n_col):
        chunks.extend(cols[i])
    # for i in range(len(chunks)):
    #     if (chunks[i]['start_row'],chunks[i]['start_col'])
    # print(chunks)
    return chunks


def generatePNG(image, chunks, tablePos):
    rowDict = {}
    colDict = {}
    for chunk in chunks:
        if chunk['end_row'] not in rowDict:
            rowDict[chunk['end_row']] = [chunk]
        else:
            rowDict[chunk['end_row']].append(chunk)
        if chunk['end_col'] not in colDict:
            colDict[chunk['end_col']] = [chunk]
        else:
            colDict[chunk['end_col']].append(chunk)
    rowLine = {}
    colLine = {}
    for row in rowDict:
        r1 = min(rowDict[row], key=lambda x: x['pos'][0])['pos'][0]
        r2 = min(rowDict[row], key=lambda x: x['pos'][1])['pos'][1]
        r3 = max(rowDict[row], key=lambda x: x['pos'][2])['pos'][2]
        r4 = max(rowDict[row], key=lambda x: x['pos'][3])['pos'][3]
        rowLine[int(row)] = [r1, r2, r3, r4]
    for col in colDict:
        c1 = min(colDict[col], key=lambda x: x['pos'][0])['pos'][0]
        c2 = min(colDict[col], key=lambda x: x['pos'][1])['pos'][1]
        c3 = max(colDict[col], key=lambda x: x['pos'][2])['pos'][2]
        c4 = max(colDict[col], key=lambda x: x['pos'][3])['pos'][3]
        colLine[int(col)] = [c1, c2, c3, c4]
    rowLine = sorted(rowLine.items(), key=lambda x: x[0])
    colLine = sorted(colLine.items(), key=lambda x: x[0])
    (x1, y1, x2, y2) = tablePos
    # draw row line
    for r in range(1, len(rowLine)):
        y = int((rowLine[r - 1][1][3] + rowLine[r][1][1]))
        cv2.line(image, (x1, y), (x2, y), (222,156,83), 2)

    # draw column line
    for c in range(1, len(colLine)):
        x = int((colLine[c - 1][1][2] + colLine[c][1][0]))
        cv2.line(image, (x, y1), (x, y2), (222,156,83), 2)
    return image


def transformStructureToTable(data, filename, savePath='./output'):
    xlsxBook = Workbook()
    workSheet = xlsxBook.active
    for cell in data:
        sr = cell['start_row'] + 1
        er = cell['end_row'] + 1
        sc = cell['start_col'] + 1
        ec = cell['end_col'] + 1
        sc_t = get_column_letter(sc)
        ec_t = get_column_letter(ec)
        pos1 = sc_t + str(sr)
        pos2 = ec_t + str(er)
        workSheet.merge_cells('{}:{}'.format(pos1, pos2))
        content = cell['text']
        workSheet.cell(sr, sc).value = content

    saveXlsx = os.path.join(savePath, filename + '.xlsx')
    xlsxBook.save(saveXlsx)


def dataInput(tuple):
    return {"pos": [tuple[0], tuple[1], tuple[2], tuple[3]], "text": tuple[4]}


def saveChunks(chunks, chunkspath, xlsxname):
    data = {}
    for chunk in chunks:
        key = '{}|{}'.format(int(chunk['pos'][0]), int(chunk['pos'][1]))
        data[key] = {
            'text': chunk['text'],
            'start_row': chunk['start_row'],
            'end_row': chunk['end_row'],
            'start_col': chunk['start_col'],
            'end_col': chunk['end_col']
        }
    with open('{}/{}.json'.format(chunkspath, xlsxname), 'w',
              encoding='utf-8') as f:
        json.dump(data, f)