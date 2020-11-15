import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def transformStructureToTable(filename, savePath='./'):
    with open(filename + '.json', encoding="utf-8") as f:
        data = json.load(f)['cells']
    xlsxBook = Workbook()
    workSheet = xlsxBook.active
    for cell in data:
        sr = cell['start_row'] + 1
        er = cell['end_row'] + 1
        sc = cell['start_col'] + 1
        ec = cell['end_col'] + 1
        sc_t = get_column_letter(sc)
        ec_t = get_column_letter(ec)
        pos1 = sc_t + str(sr)
        pos2 = ec_t + str(er)
        workSheet.merge_cells('{}:{}'.format(pos1, pos2))
        content = ' '.join(cell['content'])
        workSheet.cell(sr, sc).value = content
    saveXlsx = savePath + filename + '.xlsx'
    xlsxBook.save(saveXlsx)


if __name__ == '__main__':
    filename = 'result'
    transformStructureToTable(filename)
